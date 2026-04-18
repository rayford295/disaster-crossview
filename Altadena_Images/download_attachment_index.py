from __future__ import annotations

import argparse
import csv
import mimetypes
import re
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from threading import local
from typing import Iterable
from urllib.parse import unquote, urlparse

import openpyxl
import requests
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


URL_HEADER_CANDIDATES = {
    "url",
    "downloadurl",
    "attachmenturl",
    "link",
    "href",
}
LATITUDE_HEADER_CANDIDATES = ("lat", "latitude", "y", "ycoord", "ycoordinate")
LONGITUDE_HEADER_CANDIDATES = ("lon", "lng", "long", "longitude", "x", "xcoord", "xcoordinate")
UNSAFE_CHARS = re.compile(r"[^A-Za-z0-9._-]+")
THREAD_LOCAL = local()


@dataclass
class AttachmentRow:
    row_number: int
    source_sheet: str
    values: dict[str, str]
    url: str
    sample_id: str
    street_view_filename: str
    relative_path: Path


@dataclass
class DownloadResult:
    sample_id: str
    row_number: int
    source_sheet: str
    url: str
    relative_path: str
    status: str
    bytes_written: int
    message: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Download street-view images from a CSV/XLSX index into a dataset-ready "
            "sample folder layout, and optionally extract embedded images from Excel "
            "workbooks."
        )
    )
    parser.add_argument("index_file", help="Path to the CSV or XLSX attachment index.")
    parser.add_argument(
        "--output-root",
        help=(
            "Folder where the script writes the dataset, manifests, reports, and any "
            "extracted Excel images. Defaults to <index_stem>_output beside the index file."
        ),
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=8,
        help="Number of concurrent downloads to run. Default: 8.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        help="Per-request timeout in seconds. Default: 60.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Optional cap on the number of rows to process. Handy for testing.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Redownload files even when the target file already exists.",
    )
    parser.add_argument(
        "--skip-downloads",
        action="store_true",
        help="Skip attachment downloads and only extract embedded Excel images.",
    )
    parser.add_argument(
        "--skip-excel-images",
        action="store_true",
        help="Skip extracting embedded images from XLSX files.",
    )
    return parser.parse_args()


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9]+", "", text.strip().lower())


def sanitize_part(value: object, fallback: str = "item") -> str:
    text = "" if value is None else str(value).strip()
    text = UNSAFE_CHARS.sub("_", text).strip("._")
    return text or fallback


def sanitize_filename(filename: str, fallback_stem: str = "file") -> str:
    path = Path(filename)
    stem = sanitize_part(path.stem, fallback_stem)
    suffix = sanitize_part(path.suffix.lstrip("."), "").lower()
    return f"{stem}.{suffix}" if suffix else stem


def ensure_unique_path(relative_path: Path, seen_paths: dict[str, int]) -> Path:
    key = str(relative_path).lower()
    current = seen_paths.get(key, 0)
    seen_paths[key] = current + 1
    if current == 0:
        return relative_path

    stem = relative_path.stem
    suffix = relative_path.suffix
    updated_name = f"{stem}_{current}{suffix}"
    return relative_path.with_name(updated_name)


def normalize_extension(extension: str) -> str:
    extension = extension.strip().lower()
    if not extension:
        return ".jpg"
    if not extension.startswith("."):
        extension = f".{extension}"
    if extension in {".jpeg", ".jpe"}:
        return ".jpg"
    return extension


def pick_url_header(headers: list[str], sample_rows: Iterable[dict[str, str]]) -> str:
    normalized = {normalize_header(header): header for header in headers if header}
    for candidate in URL_HEADER_CANDIDATES:
        if candidate in normalized:
            return normalized[candidate]

    for header in headers:
        if not header:
            continue
        if any(str(row.get(header, "")).strip().startswith(("http://", "https://")) for row in sample_rows):
            return header

    raise ValueError("No URL column was found in the attachment index.")


def guess_extension(row: dict[str, str], url: str) -> str:
    name = row.get("name", "").strip()
    if name and Path(name).suffix:
        return Path(name).suffix

    file_hint = row.get("file", "").strip()
    if file_hint and Path(file_hint).suffix:
        return Path(file_hint).suffix

    parsed = Path(unquote(urlparse(url).path))
    if parsed.suffix:
        return parsed.suffix

    content_type = row.get("contentType", "").strip()
    if content_type:
        guessed = mimetypes.guess_extension(content_type)
        if guessed:
            return ".jpg" if guessed == ".jpe" else guessed

    return ""


def derive_relative_parts(file_hint: str, row: dict[str, str]) -> list[str]:
    parts = [part for part in PurePosixPath(file_hint).parts if part not in {"/", "\\"}]
    if not parts:
        return []

    fire_name = row.get("fire", "").strip()
    if fire_name and fire_name in parts:
        return parts[parts.index(fire_name) :]

    if "attachments" in parts:
        anchor = max(parts.index("attachments") - 1, 0)
        return parts[anchor:]

    if len(parts) >= 2:
        return parts[-2:]

    return parts


def build_relative_path(row: dict[str, str], url: str) -> Path:
    file_hint = row.get("file", "").strip()
    if file_hint:
        parts = derive_relative_parts(file_hint, row)
        if parts:
            sanitized_parts = [sanitize_part(part) for part in parts[:-1]]
            filename = sanitize_filename(parts[-1], "attachment")
            return Path(*sanitized_parts, filename)

    fire_name = sanitize_part(row.get("fire", ""), "attachments")
    category = sanitize_part(row.get("category", ""), "uncategorized")
    object_id = sanitize_part(row.get("objectid", ""), "object")
    attachment_id = sanitize_part(row.get("attachment_id", ""), "attachment")
    base_name = row.get("name", "").strip() or f"{object_id}_{attachment_id}{guess_extension(row, url)}"
    filename = sanitize_filename(base_name, f"{object_id}_{attachment_id}")
    if not Path(filename).suffix:
        filename = f"{filename}{guess_extension(row, url)}"
    return Path(fire_name, category, filename)


def load_csv_rows(index_path: Path) -> list[AttachmentRow]:
    with index_path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        raw_rows = [{key: "" if value is None else str(value) for key, value in row.items()} for row in reader]

    if not headers:
        raise ValueError("The CSV file does not contain a header row.")

    url_header = pick_url_header(headers, raw_rows[:25])
    rows: list[AttachmentRow] = []

    for row_number, values in enumerate(raw_rows, start=2):
        url = values.get(url_header, "").strip()
        if not url:
            continue
        rows.append(
            AttachmentRow(
                row_number=row_number,
                source_sheet="csv",
                values=values,
                url=url,
                sample_id="",
                street_view_filename="",
                relative_path=Path(),
            )
        )
    return rows


def sheet_rows_to_dicts(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[tuple[int, dict[str, str]]]:
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []

    headers = ["" if cell is None else str(cell).strip() for cell in header_row]
    rows: list[tuple[int, dict[str, str]]] = []
    for row_number, row_values in enumerate(iterator, start=2):
        row_dict = {
            headers[index]: "" if value is None else str(value)
            for index, value in enumerate(row_values)
            if index < len(headers) and headers[index]
        }
        if any(value.strip() for value in row_dict.values()):
            rows.append((row_number, row_dict))
    return rows


def load_excel_rows(index_path: Path) -> list[AttachmentRow]:
    workbook = openpyxl.load_workbook(index_path, read_only=True, data_only=True)
    try:
        collected: list[AttachmentRow] = []
        matched_sheet = False

        for sheet in workbook.worksheets:
            sheet_rows = sheet_rows_to_dicts(sheet)
            if not sheet_rows:
                continue
            headers = list(sheet_rows[0][1].keys())
            try:
                url_header = pick_url_header(headers, [row for _, row in sheet_rows[:25]])
            except ValueError:
                continue

            matched_sheet = True
            for row_number, values in sheet_rows:
                url = values.get(url_header, "").strip()
                if not url:
                    continue
                collected.append(
                    AttachmentRow(
                        row_number=row_number,
                        source_sheet=sheet.title,
                        values=values,
                        url=url,
                        sample_id="",
                        street_view_filename="",
                        relative_path=Path(),
                    )
                )

        if not matched_sheet:
            raise ValueError("No worksheet with a URL column was found in the Excel file.")

        return collected
    finally:
        workbook.close()


def validate_unique_urls(rows: list[AttachmentRow]) -> None:
    seen: dict[str, AttachmentRow] = {}
    duplicate_pairs: list[str] = []
    for row in rows:
        key = row.url.strip().lower()
        if key in seen:
            original = seen[key]
            duplicate_pairs.append(
                f"{original.source_sheet}:{original.row_number} and {row.source_sheet}:{row.row_number}"
            )
            if len(duplicate_pairs) >= 10:
                break
        else:
            seen[key] = row

    if duplicate_pairs:
        details = "; ".join(duplicate_pairs)
        raise ValueError(f"Duplicate street-view URLs were found in the index: {details}")


def assign_dataset_layout(rows: list[AttachmentRow]) -> None:
    width = max(5, len(str(len(rows))))
    for index, row in enumerate(rows, start=1):
        sample_id = f"sample_{index:0{width}d}"
        extension = normalize_extension(guess_extension(row.values, row.url))
        row.sample_id = sample_id
        row.street_view_filename = f"street_view{extension}"
        row.relative_path = Path(sample_id, row.street_view_filename)


def load_attachment_rows(index_path: Path) -> list[AttachmentRow]:
    suffix = index_path.suffix.lower()
    if suffix == ".csv":
        rows = load_csv_rows(index_path)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = load_excel_rows(index_path)
    elif suffix == ".xls":
        raise ValueError("Legacy .xls files are not supported. Please save the workbook as .xlsx first.")
    else:
        raise ValueError(f"Unsupported index format: {index_path.suffix}")

    validate_unique_urls(rows)
    assign_dataset_layout(rows)
    return rows


def build_session() -> requests.Session:
    retry = Retry(
        total=5,
        connect=5,
        read=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    session = requests.Session()
    session.headers.update({"User-Agent": "attachment-index-downloader/1.0"})
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def get_session() -> requests.Session:
    session = getattr(THREAD_LOCAL, "session", None)
    if session is None:
        session = build_session()
        THREAD_LOCAL.session = session
    return session


def download_row(
    row: AttachmentRow,
    dataset_dir: Path,
    overwrite: bool,
    timeout_seconds: int,
) -> DownloadResult:
    destination = dataset_dir / row.relative_path
    destination.parent.mkdir(parents=True, exist_ok=True)

    if destination.exists() and not overwrite:
        return DownloadResult(
            sample_id=row.sample_id,
            row_number=row.row_number,
            source_sheet=row.source_sheet,
            url=row.url,
            relative_path=str(row.relative_path),
            status="skipped",
            bytes_written=destination.stat().st_size,
            message="File already exists.",
        )

    temp_path = destination.with_suffix(destination.suffix + ".part")
    if temp_path.exists():
        temp_path.unlink()

    try:
        session = get_session()
        with session.get(row.url, stream=True, timeout=timeout_seconds) as response:
            response.raise_for_status()
            with temp_path.open("wb") as handle:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        handle.write(chunk)
        temp_path.replace(destination)
        bytes_written = destination.stat().st_size
        return DownloadResult(
            sample_id=row.sample_id,
            row_number=row.row_number,
            source_sheet=row.source_sheet,
            url=row.url,
            relative_path=str(row.relative_path),
            status="downloaded",
            bytes_written=bytes_written,
            message="OK",
        )
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink()
        return DownloadResult(
            sample_id=row.sample_id,
            row_number=row.row_number,
            source_sheet=row.source_sheet,
            url=row.url,
            relative_path=str(row.relative_path),
            status="failed",
            bytes_written=0,
            message=str(exc),
        )


def write_report(results: list[DownloadResult], report_path: Path) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "sample_id",
                "row_number",
                "source_sheet",
                "status",
                "bytes_written",
                "relative_path",
                "url",
                "message",
            ],
        )
        writer.writeheader()
        for result in results:
            writer.writerow(
                {
                    "sample_id": result.sample_id,
                    "row_number": result.row_number,
                    "source_sheet": result.source_sheet,
                    "status": result.status,
                    "bytes_written": result.bytes_written,
                    "relative_path": result.relative_path,
                    "url": result.url,
                    "message": result.message,
                }
            )


def download_rows(
    rows: list[AttachmentRow],
    dataset_dir: Path,
    workers: int,
    overwrite: bool,
    timeout_seconds: int,
) -> list[DownloadResult]:
    if not rows:
        return []

    results: list[DownloadResult] = []
    completed = 0
    total = len(rows)

    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        futures = {
            executor.submit(download_row, row, dataset_dir, overwrite, timeout_seconds): row
            for row in rows
        }
        for future in as_completed(futures):
            completed += 1
            result = future.result()
            results.append(result)
            if result.status == "failed" or completed == total or completed % 250 == 0:
                print(
                    f"[{completed}/{total}] {result.status.upper()} "
                    f"{result.relative_path}"
                )

    return sorted(results, key=lambda item: (item.source_sheet, item.row_number))


def find_row_value(values: dict[str, str], candidates: Iterable[str]) -> str:
    normalized_values = {normalize_header(key): value for key, value in values.items()}
    for candidate in candidates:
        value = normalized_values.get(candidate, "")
        if value and str(value).strip():
            return str(value).strip()
    return ""


def build_manifest_records(
    rows: list[AttachmentRow],
    results: list[DownloadResult],
) -> list[dict[str, object]]:
    results_by_key = {(result.source_sheet, result.row_number): result for result in results}
    records: list[dict[str, object]] = []

    for row in rows:
        result = results_by_key.get((row.source_sheet, row.row_number))
        download_status = result.status if result else "not_started"
        bytes_written = result.bytes_written if result else 0
        notes: list[str] = ["Remote sensing counterpart pending."]
        if result and result.status == "failed":
            notes.insert(0, f"Street-view download failed: {result.message}")

        records.append(
            {
                "pair_id": row.sample_id,
                "sample_folder": str(Path("dataset") / row.sample_id),
                "street_view_filename": row.street_view_filename,
                "street_view_relative_path": str(Path("dataset") / row.relative_path),
                "street_view_url": row.url,
                "remote_sensing_filename": "",
                "remote_sensing_relative_path": "",
                "latitude": find_row_value(row.values, LATITUDE_HEADER_CANDIDATES),
                "longitude": find_row_value(row.values, LONGITUDE_HEADER_CANDIDATES),
                "fire": row.values.get("fire", "").strip(),
                "category": row.values.get("category", "").strip(),
                "objectid": row.values.get("objectid", "").strip(),
                "attachment_id": row.values.get("attachment_id", "").strip(),
                "source_sheet": row.source_sheet,
                "source_row_number": row.row_number,
                "download_status": download_status,
                "bytes_written": bytes_written,
                "notes": " ".join(notes),
            }
        )

    return records


def write_manifest_csv(records: list[dict[str, object]], output_path: Path) -> None:
    if not records:
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(records[0].keys())
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def write_manifest_workbook(records: list[dict[str, object]], output_path: Path) -> None:
    if not records:
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "dataset_index"

    headers = list(records[0].keys())
    sheet.append(headers)
    for record in records:
        sheet.append([record[header] for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, header in enumerate(headers, start=1):
        values = [header] + ["" if row[header] is None else str(row[header]) for row in records]
        width = min(max(len(value) for value in values) + 2, 60)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    workbook.save(output_path)


def extract_excel_images(index_path: Path, output_dir: Path) -> list[str]:
    if index_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return []

    output_dir.mkdir(parents=True, exist_ok=True)
    extracted_files: list[str] = []
    seen_names: dict[str, int] = {}

    with zipfile.ZipFile(index_path, "r") as workbook_zip:
        media_entries = sorted(
            name for name in workbook_zip.namelist() if name.startswith("xl/media/") and not name.endswith("/")
        )
        for entry in media_entries:
            original_name = Path(entry).name
            safe_name = sanitize_filename(original_name, "image")
            safe_path = ensure_unique_path(Path(safe_name), seen_names)
            destination = output_dir / safe_path
            with workbook_zip.open(entry) as source, destination.open("wb") as target:
                target.write(source.read())
            extracted_files.append(str(destination))

    return extracted_files


def main() -> int:
    args = parse_args()
    index_path = Path(args.index_file).expanduser().resolve()
    if not index_path.exists():
        raise FileNotFoundError(f"Index file not found: {index_path}")

    output_root = (
        Path(args.output_root).expanduser().resolve()
        if args.output_root
        else index_path.parent / f"{index_path.stem}_output"
    )
    dataset_dir = output_root / "dataset"
    excel_images_dir = output_root / "excel_images"
    report_path = output_root / "download_report.csv"
    manifest_csv_path = output_root / "dataset_index.csv"
    manifest_xlsx_path = output_root / "dataset_index.xlsx"

    rows = load_attachment_rows(index_path)
    if args.limit:
        rows = rows[: args.limit]
        assign_dataset_layout(rows)

    print(f"Index file: {index_path}")
    print(f"Output root: {output_root}")
    print(f"Rows ready: {len(rows)}")
    print(f"Dataset dir: {dataset_dir}")

    results: list[DownloadResult] = []
    if not args.skip_downloads:
        results = download_rows(
            rows=rows,
            dataset_dir=dataset_dir,
            workers=args.workers,
            overwrite=args.overwrite,
            timeout_seconds=args.timeout,
        )
        write_report(results, report_path)
        downloaded = sum(1 for item in results if item.status == "downloaded")
        skipped = sum(1 for item in results if item.status == "skipped")
        failed = sum(1 for item in results if item.status == "failed")
        print(
            f"Download summary: downloaded={downloaded}, skipped={skipped}, "
            f"failed={failed}"
        )
        print(f"Report written to: {report_path}")
    else:
        print("Download step skipped.")

    manifest_records = build_manifest_records(rows, results)
    write_manifest_csv(manifest_records, manifest_csv_path)
    write_manifest_workbook(manifest_records, manifest_xlsx_path)
    print(f"Dataset manifest CSV: {manifest_csv_path}")
    print(f"Dataset manifest Excel: {manifest_xlsx_path}")

    if not args.skip_excel_images:
        extracted = extract_excel_images(index_path, excel_images_dir)
        if extracted:
            print(f"Extracted {len(extracted)} embedded Excel images to: {excel_images_dir}")
        elif index_path.suffix.lower() in {".xlsx", ".xlsm"}:
            print("No embedded Excel images were found in the workbook.")
    else:
        print("Excel image extraction skipped.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
