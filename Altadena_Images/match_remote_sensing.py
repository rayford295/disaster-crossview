from __future__ import annotations

import argparse
import csv
import math
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image


Image.MAX_IMAGE_PIXELS = None
PENDING_NOTE = "Remote sensing counterpart pending."


@dataclass(frozen=True)
class TileInfo:
    path: Path
    width: int
    height: int
    lon0: float
    lat0: float
    pixel_width: float
    pixel_height: float
    min_lon: float
    max_lon: float
    min_lat: float
    max_lat: float


@dataclass
class MatchCandidate:
    tile: TileInfo
    pixel_x: float
    pixel_y: float
    inside: bool
    margin_pixels: float
    distance_to_coverage_m: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Match dataset samples to georeferenced satellite GeoTIFFs and write a "
            "remote_sensing crop into each sample folder."
        )
    )
    parser.add_argument(
        "manifest_path",
        help="Path to the existing dataset manifest CSV or XLSX.",
    )
    parser.add_argument(
        "--satellite-dir",
        default="Altadena_Images",
        help="Folder containing the georeferenced .tif satellite tiles.",
    )
    parser.add_argument(
        "--dataset-root",
        help=(
            "Root dataset folder that contains sample folders. Defaults to a "
            "'dataset' folder beside the manifest."
        ),
    )
    parser.add_argument(
        "--crop-size",
        type=int,
        default=512,
        help="Square crop size in pixels for remote_sensing images. Default: 512.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Optional cap on the number of manifest rows to process for testing.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Rewrite remote_sensing images even when they already exist.",
    )
    parser.add_argument(
        "--fallback-nearest",
        action="store_true",
        help=(
            "If a point falls outside all tiles, use the nearest tile anyway. "
            "Default behavior is to leave those samples unmatched."
        ),
    )
    return parser.parse_args()


def load_manifest_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        rows = [{key: "" if value is None else str(value) for key, value in row.items()} for row in reader]
    if not headers:
        raise ValueError(f"Manifest CSV has no header row: {path}")
    return rows, headers


def load_manifest_xlsx(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise ValueError(f"Manifest workbook is empty: {path}") from exc

        headers = ["" if cell is None else str(cell).strip() for cell in header_row]
        rows: list[dict[str, str]] = []
        for row_values in iterator:
            row = {
                headers[index]: "" if value is None else str(value)
                for index, value in enumerate(row_values)
                if index < len(headers) and headers[index]
            }
            if any(value.strip() for value in row.values()):
                rows.append(row)
        if not headers:
            raise ValueError(f"Manifest workbook has no headers: {path}")
        return rows, headers
    finally:
        workbook.close()


def load_manifest(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_manifest_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_manifest_xlsx(path)
    raise ValueError(f"Unsupported manifest format: {path.suffix}")


def write_manifest_csv(records: list[dict[str, str]], headers: list[str], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(records)


def write_manifest_xlsx(records: list[dict[str, str]], headers: list[str], path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "dataset_index"
    sheet.append(headers)
    for record in records:
        sheet.append([record.get(header, "") for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, header in enumerate(headers, start=1):
        values = [header] + ["" if record.get(header) is None else str(record.get(header, "")) for record in records]
        width = min(max(len(value) for value in values) + 2, 60)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def cleaned_note(note: str) -> str:
    note = note.strip()
    if not note:
        return ""
    if note == PENDING_NOTE:
        return ""
    return note.replace(PENDING_NOTE, "").strip()


def merge_notes(*notes: str) -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for note in notes:
        cleaned = cleaned_note(note)
        if cleaned and cleaned not in seen:
            parts.append(cleaned)
            seen.add(cleaned)
    return " ".join(parts)


def first_present(row: dict[str, str], names: Iterable[str]) -> str:
    normalized = {key.strip().lower(): value for key, value in row.items()}
    for name in names:
        value = normalized.get(name.lower(), "")
        if value and str(value).strip():
            return str(value).strip()
    return ""


def load_tile_info(path: Path) -> TileInfo:
    with Image.open(path) as image:
        width, height = image.size
        tags = image.tag_v2
        scale = tags.get(33550)
        tiepoint = tags.get(33922)

    if not scale or not tiepoint:
        raise ValueError(f"GeoTIFF is missing georeferencing tags: {path.name}")

    lon0 = float(tiepoint[3])
    lat0 = float(tiepoint[4])
    pixel_width = float(scale[0])
    pixel_height = float(scale[1])
    max_lon = lon0 + pixel_width * width
    min_lat = lat0 - pixel_height * height

    return TileInfo(
        path=path,
        width=width,
        height=height,
        lon0=lon0,
        lat0=lat0,
        pixel_width=pixel_width,
        pixel_height=pixel_height,
        min_lon=lon0,
        max_lon=max_lon,
        min_lat=min_lat,
        max_lat=lat0,
    )


def load_tiles(satellite_dir: Path) -> list[TileInfo]:
    tiles = [load_tile_info(path) for path in sorted(satellite_dir.glob("*.tif"))]
    if not tiles:
        raise FileNotFoundError(f"No .tif files were found in {satellite_dir}")
    return tiles


def lon_lat_to_pixel(tile: TileInfo, lon: float, lat: float) -> tuple[float, float]:
    x = (lon - tile.lon0) / tile.pixel_width
    y = (tile.lat0 - lat) / tile.pixel_height
    return x, y


def point_inside_tile(tile: TileInfo, lon: float, lat: float) -> bool:
    return tile.min_lon <= lon <= tile.max_lon and tile.min_lat <= lat <= tile.max_lat


def margin_pixels(tile: TileInfo, pixel_x: float, pixel_y: float) -> float:
    return min(pixel_x, tile.width - 1 - pixel_x, pixel_y, tile.height - 1 - pixel_y)


def distance_to_tile_coverage_m(tile: TileInfo, lon: float, lat: float) -> float:
    dx = 0.0
    dy = 0.0
    if lon < tile.min_lon:
        dx = tile.min_lon - lon
    elif lon > tile.max_lon:
        dx = lon - tile.max_lon

    if lat < tile.min_lat:
        dy = tile.min_lat - lat
    elif lat > tile.max_lat:
        dy = lat - tile.max_lat

    lon_scale = 111320 * math.cos(math.radians(lat))
    lat_scale = 111320
    return math.hypot(dx * lon_scale, dy * lat_scale)


def best_match_for_point(
    tiles: list[TileInfo],
    lon: float,
    lat: float,
    fallback_nearest: bool,
) -> MatchCandidate | None:
    inside_candidates: list[MatchCandidate] = []
    nearest_candidates: list[MatchCandidate] = []

    for tile in tiles:
        pixel_x, pixel_y = lon_lat_to_pixel(tile, lon, lat)
        if point_inside_tile(tile, lon, lat):
            inside_candidates.append(
                MatchCandidate(
                    tile=tile,
                    pixel_x=pixel_x,
                    pixel_y=pixel_y,
                    inside=True,
                    margin_pixels=margin_pixels(tile, pixel_x, pixel_y),
                    distance_to_coverage_m=0.0,
                )
            )
        else:
            nearest_candidates.append(
                MatchCandidate(
                    tile=tile,
                    pixel_x=min(max(pixel_x, 0.0), tile.width - 1),
                    pixel_y=min(max(pixel_y, 0.0), tile.height - 1),
                    inside=False,
                    margin_pixels=-1.0,
                    distance_to_coverage_m=distance_to_tile_coverage_m(tile, lon, lat),
                )
            )

    if inside_candidates:
        return max(
            inside_candidates,
            key=lambda candidate: (candidate.margin_pixels, candidate.tile.width * candidate.tile.height),
        )

    if fallback_nearest and nearest_candidates:
        return min(nearest_candidates, key=lambda candidate: candidate.distance_to_coverage_m)

    return None


def crop_box_for_pixel(tile: TileInfo, pixel_x: float, pixel_y: float, crop_size: int) -> tuple[int, int, int, int]:
    crop_size = max(1, crop_size)
    crop_width = min(crop_size, tile.width)
    crop_height = min(crop_size, tile.height)

    left = int(round(pixel_x - crop_width / 2))
    top = int(round(pixel_y - crop_height / 2))
    left = min(max(left, 0), tile.width - crop_width)
    top = min(max(top, 0), tile.height - crop_height)
    right = left + crop_width
    bottom = top + crop_height
    return left, top, right, bottom


def infer_sample_dir(row: dict[str, str], dataset_root: Path) -> Path:
    sample_folder = row.get("sample_folder", "").strip()
    pair_id = row.get("pair_id", "").strip()

    if sample_folder:
        return dataset_root.parent / Path(sample_folder)
    if pair_id:
        return dataset_root / pair_id
    raise ValueError("Manifest row is missing both 'sample_folder' and 'pair_id'.")


def with_added_headers(headers: list[str], extras: Iterable[str]) -> list[str]:
    updated = list(headers)
    for extra in extras:
        if extra not in updated:
            updated.append(extra)
    return updated


def process_manifest_rows(
    rows: list[dict[str, str]],
    tiles: list[TileInfo],
    dataset_root: Path,
    crop_size: int,
    overwrite: bool,
    fallback_nearest: bool,
) -> list[dict[str, str]]:
    processed_rows: list[dict[str, str]] = []
    jobs_by_tile: dict[Path, list[tuple[dict[str, str], tuple[int, int, int, int], Path]]] = defaultdict(list)

    for row in rows:
        updated = dict(row)
        latitude_text = first_present(updated, ("latitude", "lat"))
        longitude_text = first_present(updated, ("longitude", "lon", "lng", "long"))
        if not latitude_text or not longitude_text:
            updated["remote_match_status"] = "missing_coordinates"
            updated["remote_match_notes"] = merge_notes(
                updated.get("remote_match_notes", ""),
                "Missing latitude/longitude in manifest.",
            )
            updated["notes"] = merge_notes(updated.get("notes", ""), updated["remote_match_notes"])
            processed_rows.append(updated)
            continue

        lat = float(latitude_text)
        lon = float(longitude_text)
        match = best_match_for_point(tiles, lon, lat, fallback_nearest)
        sample_dir = infer_sample_dir(updated, dataset_root)
        sample_dir.mkdir(parents=True, exist_ok=True)
        remote_filename = "remote_sensing.jpg"
        remote_path = sample_dir / remote_filename
        street_view_path = sample_dir / updated.get("street_view_filename", "street_view.jpg")

        if match is None:
            updated["remote_sensing_filename"] = ""
            updated["remote_sensing_relative_path"] = ""
            updated["remote_match_status"] = "unmatched"
            updated["remote_tile_filename"] = ""
            updated["remote_tile_path"] = ""
            updated["remote_tile_aux_xml_path"] = ""
            updated["remote_pixel_x"] = ""
            updated["remote_pixel_y"] = ""
            updated["remote_crop_box"] = ""
            updated["remote_distance_to_coverage_m"] = ""
            updated["remote_match_notes"] = merge_notes(
                updated.get("remote_match_notes", ""),
                "No satellite tile covers this coordinate.",
                "Street-view pair remains incomplete.",
            )
            updated["notes"] = merge_notes(updated.get("notes", ""), updated["remote_match_notes"])
            processed_rows.append(updated)
            continue

        crop_box = crop_box_for_pixel(match.tile, match.pixel_x, match.pixel_y, crop_size)
        relative_remote_path = Path("dataset") / sample_dir.name / remote_filename
        remote_status = "matched" if match.inside else "nearest_tile"
        if remote_path.exists() and not overwrite:
            remote_status = "matched_existing" if match.inside else "nearest_tile_existing"

        notes: list[str] = []
        aux_xml_path = match.tile.path.with_suffix(match.tile.path.suffix + ".aux.xml")
        if not match.inside:
            notes.append(
                f"Coordinate is outside imagery coverage; nearest tile used at {match.distance_to_coverage_m:.1f} m."
            )
        if not street_view_path.exists():
            notes.append("Street-view image is not present in the sample folder yet.")

        updated["remote_sensing_filename"] = remote_filename
        updated["remote_sensing_relative_path"] = str(relative_remote_path)
        updated["remote_match_status"] = remote_status
        updated["remote_tile_filename"] = match.tile.path.name
        updated["remote_tile_path"] = str(match.tile.path)
        updated["remote_tile_aux_xml_path"] = str(aux_xml_path) if aux_xml_path.exists() else ""
        updated["remote_pixel_x"] = f"{match.pixel_x:.2f}"
        updated["remote_pixel_y"] = f"{match.pixel_y:.2f}"
        updated["remote_crop_box"] = ",".join(str(value) for value in crop_box)
        updated["remote_distance_to_coverage_m"] = (
            f"{match.distance_to_coverage_m:.2f}" if not match.inside else "0.00"
        )
        updated["remote_match_notes"] = merge_notes(updated.get("remote_match_notes", ""), *notes)
        updated["notes"] = merge_notes(updated.get("notes", ""), updated["remote_match_notes"])
        processed_rows.append(updated)

        if not remote_path.exists() or overwrite:
            jobs_by_tile[match.tile.path].append((updated, crop_box, remote_path))

    for tile_path, jobs in jobs_by_tile.items():
        with Image.open(tile_path) as image:
            for row, crop_box, remote_path in jobs:
                crop = image.crop(crop_box)
                if crop.mode != "RGB":
                    crop = crop.convert("RGB")
                crop.save(remote_path, quality=95)

    return processed_rows


def summarize(rows: list[dict[str, str]]) -> dict[str, int]:
    summary: dict[str, int] = defaultdict(int)
    for row in rows:
        summary[row.get("remote_match_status", "unknown")] += 1
    return dict(sorted(summary.items()))


def output_paths(manifest_path: Path) -> tuple[Path, Path]:
    base = manifest_path.with_suffix("")
    return (
        base.with_suffix(".csv"),
        base.with_suffix(".xlsx"),
    )


def main() -> int:
    args = parse_args()
    manifest_path = Path(args.manifest_path).expanduser().resolve()
    if not manifest_path.exists():
        raise FileNotFoundError(f"Manifest not found: {manifest_path}")

    satellite_dir = Path(args.satellite_dir).expanduser().resolve()
    dataset_root = (
        Path(args.dataset_root).expanduser().resolve()
        if args.dataset_root
        else manifest_path.parent / "dataset"
    )
    if not satellite_dir.exists():
        raise FileNotFoundError(f"Satellite directory not found: {satellite_dir}")

    rows, headers = load_manifest(manifest_path)
    if args.limit:
        rows = rows[: args.limit]
    tiles = load_tiles(satellite_dir)

    print(f"Manifest: {manifest_path}")
    print(f"Dataset root: {dataset_root}")
    print(f"Satellite dir: {satellite_dir}")
    print(f"Rows to process: {len(rows)}")
    print(f"Satellite tiles: {len(tiles)}")

    processed_rows = process_manifest_rows(
        rows=rows,
        tiles=tiles,
        dataset_root=dataset_root,
        crop_size=args.crop_size,
        overwrite=args.overwrite,
        fallback_nearest=args.fallback_nearest,
    )

    headers = with_added_headers(
        headers,
        (
            "remote_tile_filename",
            "remote_tile_path",
            "remote_tile_aux_xml_path",
            "remote_match_status",
            "remote_match_notes",
            "remote_pixel_x",
            "remote_pixel_y",
            "remote_crop_box",
            "remote_distance_to_coverage_m",
        ),
    )
    csv_path, xlsx_path = output_paths(manifest_path)
    write_manifest_csv(processed_rows, headers, csv_path)
    write_manifest_xlsx(processed_rows, headers, xlsx_path)

    print(f"Updated manifest CSV: {csv_path}")
    print(f"Updated manifest Excel: {xlsx_path}")
    print("Summary:")
    for status, count in summarize(processed_rows).items():
        print(f"  {status}: {count}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
